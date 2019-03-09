# coding=UTF-8
import openpyxl


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def append_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.load_workbook(path)
    # 切换到目标数据表
    # ws = wb[]
    sheet = workbook[sheet_name]
    for x in value3:
        sheet.append(x)
    savename = 'test.xlsx'
    workbook.save(savename)


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


book_name_xlsx = 'test.xlsx'

sheet_name_xlsx = 'xlsx格式测试表'

value3 = [["姓名", "性别", "年龄", "城市", "职业"],
          ["111", "女", "66", "石家庄", "运维工程师"],
          ["222", "男", "55", "南京", "饭店老板"],
          ["333", "女", "27", "苏州", "保安"], ]

# write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
# read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)

if __name__ == '__main__':
    # write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
    # read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)
    append_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
