# coding=UTF-8
import openpyxl


class Read_write:

    def creat_excel_file(self, path):
        workbook = openpyxl.Workbook()
        workbook.save(path)

    # def create_sheet(self,file_path,sheet_name):
    #     workbook = openpyxl.load_workbook(file_path)
    #     workbook.create_sheet(title=sheet_name)
    #     workbook.save(file_path)

    def create_sheet(self, file_path, sheet_name, title_value):
        index = len(title_value)
        workbook = openpyxl.load_workbook(file_path)
        workbook.create_sheet(title=sheet_name)
        sheet = workbook[sheet_name]
        for i in range(0, index):
            sheet.cell(1, column=i + 1, value=str(title_value[i]))
        workbook.save(file_path)

    # def initial_sheet_title(self,file_path,sheet_name,value):
    #     index = len(value)
    #     workbook= openpyxl.Workbook()
    #     sheet = workbook.active
    #     sheet.title = sheet_name
    #     for i in range(0, index):
    #         sheet.cell(1, column=i + 1, value=str(value[i]))
    #     workbook.save(file_path)

    # def creat_excel_file(self, path, sheet_name, value):
    #     index = len(value)
    #     workbook = openpyxl.Workbook()
    #     sheet = workbook.active
    #     # sheet.title = sheet_name
    #     # for i in range(0, index):
    #     #     sheet.cell(1, column=i + 1, value=str(value[i]))
    #     workbook.save(path)

    def append_excel_xlsx(self, path, sheet_name, value):
        index = len(value)
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        sheet.append(value)
        workbook.save(path)

    def read_excel_xlsx(self, path, sheet_name):
        workbook = openpyxl.load_workbook(path)

        # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
        list_tem = []

        sheet = workbook[sheet_name]
        for row in sheet.rows:
            row_tem = []
            for cell in row:
                row_tem.append(cell.value)
                # print(cell.value, "\t", end="")
            list_tem.append(row_tem)

        return  list_tem

# if __name__ == '__main__':
#
#     Read_write.write_excel_xlsx("test.xlsx","sheet_test","888")
# print(value)
# Read_write.append_excel_xlsx(file_path,sheet_name,value)
# Read_write.write_excel_xlsx(file_path, sheet_name, value)
# read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)
# append_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
# readWrite = Read_write()
